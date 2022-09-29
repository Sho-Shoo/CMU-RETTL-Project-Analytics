# Run this python script to annotate any partial correlation table file
# Rerquirements for the table file:
# - column number must be twice of row number 
# - each column pair signifies correlation coefficient and p-value for the correlation

from operator import index
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Color, PatternFill, Font, Border
from openpyxl.styles import colors
from openpyxl.cell import Cell
import csv

def createExcel(filePath, outputPath, index_col="varName"): 

    assert ".csv" in filePath
    table = pd.read_csv(filePath, index_col=index_col) 
    assert len(table) * 2 == len(table.columns) # column number must be twice of row number 

    # initialize a new openpyxl worksheet 
    wb = Workbook()
    ws = wb.active # new worksheet (ws)

    # csv 2 worksheet 
    with open(filePath) as f:
        reader = csv.reader(f, delimiter=',')
        for row in reader:
            ws.append(row)

    # access p-value columns and highlight those < 0.1 and < 0.05
    # red for < 0.05
    redFill = PatternFill(start_color='FFFF0000',
                          end_color='FFFF0000',
                          fill_type='solid')
    # yellow for < 0.1
    yellowFill = PatternFill(start_color='FFFF00',
                             end_color='FFFF00',
                             fill_type='solid')

    nrow = ws.max_row 
    ncol = ws.max_column
    for col in range(1, ncol+1): 
        # this condition indicates that it is an p value column 
        if "p_value" in ws.cell(row=1, column=col).value: 
            # if is p value column, go down the column 
            for row in range(2, nrow+1): 
                # fill different values with corresponding colors 
                if float(ws.cell(row=row, column=col).value) < 0.05: 
                    ws.cell(row=row, column=col).fill = redFill
                elif float(ws.cell(row=row, column=col).value) < 0.1: 
                    ws.cell(row=row, column=col).fill = yellowFill


    # save workbook to an output file
    wb.save(outputPath)


if __name__ == "__main__": 

    createExcel("output_files/partial_corrlation_table.csv", 
                "output_files/part_corr_tb_annotated.xlsx")


